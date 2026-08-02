"""
Company Export Service
======================
Generates an Excel workbook containing all companies for a user.
Uses the same openpyxl styling conventions as excel_service.py.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.models.company import Company


SHEET_NAME = "Companies"

COLUMNS = [
    "Company ID",
    "Name",
    "Website",
    "Email",
    "Rating",
    "Applications Count",
    "Created Date",
]


# ---------------------------------------------------------------------------
# Styling helpers (mirrors excel_service.py)
# ---------------------------------------------------------------------------

def _header_font() -> Font:
    return Font(bold=True, size=11)


def _thin_border() -> Border:
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_header_row(ws, columns: list[str]) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _header_font()
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _apply_data_cell(cell, value) -> None:
    cell.value = value
    cell.border = _thin_border()
    cell.alignment = Alignment(vertical="center", wrap_text=False)


def _autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


# ---------------------------------------------------------------------------
# Export entry point
# ---------------------------------------------------------------------------

def export_companies_to_excel(db: Session, user_id: int) -> bytes:
    """
    Build a workbook with one Companies sheet.
    Returns raw bytes ready to be streamed as a file download.
    """
    companies: list[Company] = (
        db.query(Company)
        .options(joinedload(Company.applications))
        .filter(Company.user_id == user_id)
        .order_by(Company.name.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.row_dimensions[1].height = 20

    _apply_header_row(ws, COLUMNS)

    for row_idx, company in enumerate(companies, start=2):
        # "Created Date" is not a real column on the Company model,
        # so we derive it from the earliest application date, or leave blank.
        # The model has no created_at; use the min applied_date as a proxy.
        app_dates = [a.applied_date for a in company.applications if a.applied_date]
        created_date = str(min(app_dates)) if app_dates else ""

        values = [
            company.id,
            company.name,
            company.website or "",
            company.email or "",
            company.rating.value,
            len(company.applications),
            created_date,
        ]
        for col_idx, value in enumerate(values, start=1):
            _apply_data_cell(ws.cell(row=row_idx, column=col_idx), value)

    _autosize_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
