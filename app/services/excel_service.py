"""
Excel Import / Export Service
==============================
Handles all openpyxl logic for exporting a user's job search data to an
Excel workbook and importing it back.  No HTTP concerns live here.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.models.application import JobApplication, ApplicationStatus
from app.models.followup import ApplicationFollowUp, FollowupType, FollowupResponse
from app.crud.company import get_or_create_company
from app.crud.resume import get_resume_by_display_name


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

APPLICATIONS_SHEET = "Applications"
FOLLOWUPS_SHEET = "FollowUps"

APPLICATIONS_COLUMNS = [
    "Application ID",
    "Company Name",
    "Company Website",
    "Company Email",
    "Position",
    "Application Method",
    "Job Posting URL",
    "Applied Date",
    "Next Follow-up Date",
    "Resume Name",
    "Status",
    "Notes",
    "Created At",
    "Updated At",
    "Follow-up Count",
]

FOLLOWUPS_COLUMNS = [
    "Application ID",
    "Company Name",
    "Position",
    "Follow-up Number",
    "Follow-up Date",
    "Follow-up Type",
    "Response",
    "Notes",
]

# Required columns the importer must find (subset — derived cols are excluded)
# Application ID is intentionally NOT required — blank IDs are treated as new applications.
REQUIRED_APP_COLUMNS = {
    "Company Name",
    "Position",
    "Application Method",
    "Applied Date",
    "Status",
}

REQUIRED_FOLLOWUP_COLUMNS = {
    "Application ID",
    "Follow-up Date",
    "Follow-up Type",
    "Response",
}


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def _header_font() -> Font:
    return Font(bold=True, size=11)


def _thin_border() -> Border:
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_header_row(ws, columns: list[str]) -> None:
    """Write bold header row with light border and freeze it."""
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _header_font()
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _apply_data_cell(cell, value: Any) -> None:
    cell.value = value
    cell.border = _thin_border()
    cell.alignment = Alignment(vertical="center", wrap_text=False)


def _autosize_columns(ws) -> None:
    """Set each column width to the max content length (capped at 60)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _fmt_date(d: date | datetime | None) -> str:
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%d %H:%M:%S")
    return d.strftime("%Y-%m-%d")


# ---------------------------------------------------------------------------
# EXPORT
# ---------------------------------------------------------------------------

def export_to_excel(db: Session, user_id: int) -> bytes:
    """
    Build a workbook containing all of the user's applications and follow-ups.
    Returns raw bytes ready to be streamed as a file download.
    """
    # Eager-load every relationship we will touch during template rendering
    applications: list[JobApplication] = (
        db.query(JobApplication)
        .options(
            joinedload(JobApplication.company),
            joinedload(JobApplication.followups),
        )
        .filter(JobApplication.user_id == user_id)
        .order_by(JobApplication.applied_date.desc())
        .all()
    )

    wb = Workbook()

    # --- Applications sheet ---
    ws_apps = wb.active
    ws_apps.title = APPLICATIONS_SHEET
    ws_apps.row_dimensions[1].height = 20

    _apply_header_row(ws_apps, APPLICATIONS_COLUMNS)

    for row_idx, app in enumerate(applications, start=2):
        values = [
            app.id,
            app.company.name,
            app.company.website or "",
            app.company.email or "",
            app.position,
            app.application_method,
            app.job_post_url or "",
            _fmt_date(app.applied_date),
            _fmt_date(app.next_followup_date),
            app.resume.display_name if app.resume else "",
            app.status.value,
            app.notes or "",
            _fmt_date(app.created_at),
            _fmt_date(app.updated_at),
            len(app.followups),
        ]
        for col_idx, value in enumerate(values, start=1):
            _apply_data_cell(ws_apps.cell(row=row_idx, column=col_idx), value)

    _autosize_columns(ws_apps)

    # --- FollowUps sheet ---
    ws_fu = wb.create_sheet(title=FOLLOWUPS_SHEET)
    ws_fu.row_dimensions[1].height = 20

    _apply_header_row(ws_fu, FOLLOWUPS_COLUMNS)

    fu_row = 2
    for app in applications:
        # Sort follow-ups by date ascending
        sorted_followups = sorted(app.followups, key=lambda f: f.followup_date)
        for fu_num, fu in enumerate(sorted_followups, start=1):
            values = [
                app.id,
                app.company.name,
                app.position,
                fu_num,
                _fmt_date(fu.followup_date),
                fu.followup_type.value,
                fu.response.value,
                fu.notes or "",
            ]
            for col_idx, value in enumerate(values, start=1):
                _apply_data_cell(ws_fu.cell(row=fu_row, column=col_idx), value)
            fu_row += 1

    _autosize_columns(ws_fu)

    # Serialise to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# IMPORT — validation helpers
# ---------------------------------------------------------------------------

class ImportError(Exception):
    """Raised for any workbook / data validation failure."""


def _validate_workbook_structure(wb: Workbook) -> None:
    sheet_names = wb.sheetnames
    for required in (APPLICATIONS_SHEET, FOLLOWUPS_SHEET):
        if required not in sheet_names:
            raise ImportError(
                f"Missing worksheet '{required}'. "
                f"Found sheets: {', '.join(sheet_names)}. "
                "Please export a fresh workbook from this application."
            )


def _get_column_map(ws) -> dict[str, int]:
    """Return {column_header: column_index (1-based)} from row 1."""
    mapping: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            mapping[str(cell.value).strip()] = cell.column
    return mapping


def _validate_columns(col_map: dict[str, int], required: set[str], sheet_name: str) -> None:
    missing = required - set(col_map.keys())
    if missing:
        raise ImportError(
            f"Worksheet '{sheet_name}' is missing required columns: "
            + ", ".join(sorted(missing))
        )


def _cell_value(row, col_map: dict[str, int], column_name: str) -> Any:
    """Safely read a cell value; returns None if column absent or cell empty."""
    col_idx = col_map.get(column_name)
    if col_idx is None:
        return None
    cell = row[col_idx - 1]  # row is a tuple (0-indexed)
    val = cell.value
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return val


def _parse_date(raw: Any, column_name: str, row_num: int) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, (date, datetime)):
        return raw.date() if isinstance(raw, datetime) else raw
    if isinstance(raw, str):
        raw = raw.strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        raise ImportError(
            f"Row {row_num}: '{column_name}' has an unrecognised date format: '{raw}'. "
            "Expected YYYY-MM-DD."
        )
    raise ImportError(
        f"Row {row_num}: '{column_name}' contains an unexpected value type: {type(raw).__name__}."
    )


def _parse_status(raw: Any, row_num: int) -> ApplicationStatus:
    if raw is None:
        raise ImportError(f"Row {row_num}: 'Status' is required.")
    try:
        return ApplicationStatus(str(raw).strip().upper())
    except ValueError:
        valid = ", ".join(s.value for s in ApplicationStatus)
        raise ImportError(
            f"Row {row_num}: 'Status' value '{raw}' is not valid. "
            f"Allowed values: {valid}."
        )


def _parse_followup_type(raw: Any, row_num: int) -> FollowupType:
    if raw is None:
        raise ImportError(f"Row {row_num}: 'Follow-up Type' is required.")
    raw_str = str(raw).strip()
    for member in FollowupType:
        if member.value.lower() == raw_str.lower():
            return member
    valid = ", ".join(m.value for m in FollowupType)
    raise ImportError(
        f"Row {row_num}: 'Follow-up Type' value '{raw_str}' is not valid. "
        f"Allowed values: {valid}."
    )


def _parse_followup_response(raw: Any, row_num: int) -> FollowupResponse:
    if raw is None:
        raise ImportError(f"Row {row_num}: 'Response' is required.")
    raw_str = str(raw).strip()
    for member in FollowupResponse:
        if member.value.lower() == raw_str.lower():
            return member
    valid = ", ".join(m.value for m in FollowupResponse)
    raise ImportError(
        f"Row {row_num}: 'Response' value '{raw_str}' is not valid. "
        f"Allowed values: {valid}."
    )


# ---------------------------------------------------------------------------
# IMPORT — main entry point
# ---------------------------------------------------------------------------

def import_from_excel(db: Session, user_id: int, file_bytes: bytes) -> dict:
    """
    Parse the workbook, validate every row, then import everything inside a
    single database transaction.  Raises ImportError on any problem so the
    caller can surface a friendly message without a partial commit.

    Returns a summary dict on success: {"applications": int, "followups": int}
    """
    # --- Load workbook ---
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportError(f"Could not open workbook: {exc}. The file may be corrupted.")

    # --- Structural validation ---
    _validate_workbook_structure(wb)

    ws_apps = wb[APPLICATIONS_SHEET]
    ws_fu = wb[FOLLOWUPS_SHEET]

    # Check for completely empty sheets (only header row)
    app_rows = list(ws_apps.iter_rows(min_row=2, values_only=False))
    # Filter out entirely blank rows
    app_rows = [r for r in app_rows if any(c.value is not None for c in r)]

    if not app_rows:
        raise ImportError(
            f"Worksheet '{APPLICATIONS_SHEET}' contains no data rows. "
            "Nothing to import."
        )

    fu_rows = list(ws_fu.iter_rows(min_row=2, values_only=False))
    fu_rows = [r for r in fu_rows if any(c.value is not None for c in r)]

    # --- Column maps ---
    app_col = _get_column_map(ws_apps)
    fu_col = _get_column_map(ws_fu)

    _validate_columns(app_col, REQUIRED_APP_COLUMNS, APPLICATIONS_SHEET)
    _validate_columns(fu_col, REQUIRED_FOLLOWUP_COLUMNS, FOLLOWUPS_SHEET)

    # --- Parse Applications rows ---
    # row_key: the value used internally to link follow-ups to applications.
    # - If "Application ID" is supplied and is a valid integer  → use it as the key.
    # - If "Application ID" is blank                            → use -(row_num) as a
    #   synthetic negative key (never written to the DB; PG generates the real PK).
    # Duplicate supplied IDs within the same import file are still an error.
    parsed_apps: list[dict] = []
    seen_supplied_ids: set[int] = set()

    for row_num, row in enumerate(app_rows, start=2):
        original_id_raw = _cell_value(row, app_col, "Application ID")
        company_name = _cell_value(row, app_col, "Company Name")
        position = _cell_value(row, app_col, "Position")
        application_method = _cell_value(row, app_col, "Application Method")
        applied_date_raw = _cell_value(row, app_col, "Applied Date")

        if not company_name:
            raise ImportError(f"Row {row_num} in '{APPLICATIONS_SHEET}': 'Company Name' is required.")
        if not position:
            raise ImportError(f"Row {row_num} in '{APPLICATIONS_SHEET}': 'Position' is required.")
        if not application_method:
            raise ImportError(f"Row {row_num} in '{APPLICATIONS_SHEET}': 'Application Method' is required.")

        # --- Application ID: optional ---
        if original_id_raw is None:
            # Blank — assign synthetic internal key (negative, never touches DB)
            row_key = -row_num
        else:
            try:
                row_key = int(original_id_raw)
            except (ValueError, TypeError):
                raise ImportError(
                    f"Row {row_num} in '{APPLICATIONS_SHEET}': "
                    f"'Application ID' must be a number or blank, got '{original_id_raw}'."
                )
            # Only check duplicates for explicitly supplied IDs
            if row_key in seen_supplied_ids:
                raise ImportError(
                    f"Row {row_num} in '{APPLICATIONS_SHEET}': "
                    f"Duplicate 'Application ID' {row_key}. Each supplied ID must be unique."
                )
            seen_supplied_ids.add(row_key)

        applied_date = _parse_date(applied_date_raw, "Applied Date", row_num)
        if applied_date is None:
            raise ImportError(f"Row {row_num} in '{APPLICATIONS_SHEET}': 'Applied Date' is required.")

        next_followup_raw = _cell_value(row, app_col, "Next Follow-up Date")
        next_followup_date = _parse_date(next_followup_raw, "Next Follow-up Date", row_num)

        status = _parse_status(_cell_value(row, app_col, "Status"), row_num)

        parsed_apps.append({
            "row_key": row_key,          # internal mapping key only, never written to DB
            "company_name": company_name,
            "company_website": _cell_value(row, app_col, "Company Website"),
            "company_email": _cell_value(row, app_col, "Company Email"),
            "position": position,
            "application_method": application_method,
            "job_post_url": _cell_value(row, app_col, "Job Posting URL"),
            "applied_date": applied_date,
            "next_followup_date": next_followup_date,
            "resume_name": _cell_value(row, app_col, "Resume Name"),
            "status": status,
            "notes": _cell_value(row, app_col, "Notes"),
        })

    # --- Parse FollowUp rows ---
    parsed_followups: list[dict] = []
    for row_num, row in enumerate(fu_rows, start=2):
        app_id_raw = _cell_value(row, fu_col, "Application ID")
        fu_date_raw = _cell_value(row, fu_col, "Follow-up Date")
        fu_type_raw = _cell_value(row, fu_col, "Follow-up Type")
        fu_response_raw = _cell_value(row, fu_col, "Response")

        # Application ID on follow-ups: blank is allowed (treated as new-app reference).
        # Must be a valid integer when supplied.
        if app_id_raw is None:
            ref_key = -row_num  # synthetic — will be resolved via id_map below
        else:
            try:
                ref_key = int(app_id_raw)
            except (ValueError, TypeError):
                raise ImportError(
                    f"Row {row_num} in '{FOLLOWUPS_SHEET}': "
                    f"'Application ID' must be a number or blank, got '{app_id_raw}'."
                )

        fu_date = _parse_date(fu_date_raw, "Follow-up Date", row_num)
        if fu_date is None:
            raise ImportError(f"Row {row_num} in '{FOLLOWUPS_SHEET}': 'Follow-up Date' is required.")

        fu_type = _parse_followup_type(fu_type_raw, row_num)
        fu_response = _parse_followup_response(fu_response_raw, row_num)

        parsed_followups.append({
            "ref_key": ref_key,
            "followup_date": fu_date,
            "followup_type": fu_type,
            "response": fu_response,
            "notes": _cell_value(row, fu_col, "Notes"),
        })

    # --- Validate follow-up references ---
    # A follow-up's ref_key must match some app's row_key in this import batch.
    row_keys = {a["row_key"] for a in parsed_apps}
    for fu in parsed_followups:
        if fu["ref_key"] not in row_keys:
            raise ImportError(
                f"FollowUps sheet references Application ID {fu['ref_key']} "
                "which does not exist in the Applications sheet."
            )

    # --- All validation passed — write to database inside a single transaction ---
    # Maps row_key (internal) -> newly created JobApplication (with real PG-generated PK)
    id_map: dict[int, JobApplication] = {}

    try:
        # 1. Companies + Applications
        for app_data in parsed_apps:
            company = get_or_create_company(
                db=db,
                user_id=user_id,
                name=app_data["company_name"],
                company_id=None,
                website=app_data["company_website"],
                email=app_data["company_email"],
            )
            # Update website/email only when the existing field is currently NULL/empty
            if app_data["company_website"] and not company.website:
                company.website = app_data["company_website"]
            if app_data["company_email"] and not company.email:
                company.email = app_data["company_email"]

            # Resolve resume by display name — never fail if not found
            resume_id = None
            if app_data["resume_name"]:
                matched = get_resume_by_display_name(
                    db=db,
                    user_id=user_id,
                    display_name=app_data["resume_name"],
                )
                if matched:
                    resume_id = matched.id

            new_app = JobApplication(
                position=app_data["position"],
                application_method=app_data["application_method"],
                job_post_url=app_data["job_post_url"],
                applied_date=app_data["applied_date"],
                next_followup_date=app_data["next_followup_date"],
                resume_id=resume_id,
                notes=app_data["notes"],
                status=app_data["status"],
                user_id=user_id,
                company_id=company.id,
            )
            db.add(new_app)
            db.flush()  # PostgreSQL assigns the real PK here — row_key is never written
            id_map[app_data["row_key"]] = new_app

        # 2. Follow-ups — resolved via id_map using ref_key
        for fu_data in parsed_followups:
            new_app = id_map[fu_data["ref_key"]]
            fu = ApplicationFollowUp(
                application_id=new_app.id,
                followup_date=fu_data["followup_date"],
                followup_type=fu_data["followup_type"],
                response=fu_data["response"],
                notes=fu_data["notes"],
            )
            db.add(fu)

        db.commit()

    except ImportError:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise ImportError(f"Database error during import: {exc}") from exc

    return {
        "applications": len(parsed_apps),
        "followups": len(parsed_followups),
    }
